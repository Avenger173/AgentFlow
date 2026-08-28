"""D5.3 字段加工与新副本的离线端到端回归。

所有输入和输出均位于临时目录。覆盖十类受限加工、除零与未知列拒绝、确认后数据副本回读、
任务历史脱敏、协作式取消和源文件字节不变；不会调用模型、网络或真实客户文件。
"""

from __future__ import annotations

import atexit
import asyncio
import base64
import csv
from hashlib import sha256
from io import BytesIO
import os
from pathlib import Path
import shutil
import sys
import tempfile
import time

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1]
VERIFY_ROOT = Path(tempfile.mkdtemp(prefix="agentflow_data_transform_verify_"))
os.environ["AGENTFLOW_DATA_DIR"] = str(VERIFY_ROOT / "data")
os.environ["AGENTFLOW_DATA_WORKSPACE_DIR"] = str(VERIFY_ROOT / "workspace")
os.environ["AGENTFLOW_DATA_TRANSFORMATION_OUTPUT_DIR"] = str(VERIFY_ROOT / "output" / "data_transformations")
os.environ["AGENTFLOW_CHAT_MODE"] = "mock"
atexit.register(lambda: shutil.rmtree(VERIFY_ROOT, ignore_errors=True))
sys.path.insert(0, str(BACKEND_ROOT))

from app.schemas.data_agent import DataTransformationExportRequest  # noqa: E402
from app.services.data_transformation_delivery import (  # noqa: E402
    cancel_data_transformation_task,
    create_data_transformation_queued_run,
)
from main import app  # noqa: E402


def _base64(data: bytes) -> str:
    return base64.b64encode(data).decode("ascii")


def _wait_for_terminal_result(client: TestClient, task_id: str) -> dict[str, object]:
    for _ in range(300):
        response = client.get(f"/api/agents/data_agent/transformations/export/{task_id}/result")
        assert response.status_code == 200, response.text
        payload = response.json()
        if payload["status"] not in {"pending", "running", "queued"}:
            return payload
        time.sleep(0.05)
    raise AssertionError("字段加工任务在 15 秒内没有到达终态")


def _preview(client: TestClient, request: dict[str, object]) -> dict[str, object]:
    response = client.post("/api/agents/data_agent/transformations/preview", json=request)
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["plan"]["result_column"]
    assert payload["row_count"] == 4
    return payload


def main() -> None:
    source = (
        "日期,单价,数量,成绩,姓名\n"
        "2026-01-01,12,2,92, Alice \n"
        "2026-02-01,15,0,81,Bob\n"
        "2026-03-01,20,3,59, 陈晨 \n"
        "2026-04-01,18,4,76,Dana\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        imported = client.post(
            "/api/agents/data_agent/datasets",
            json={"filename": "字段加工样本.csv", "content_base64": _base64(source)},
        )
        assert imported.status_code == 200, imported.text
        profile_response = client.get("/api/agents/data_agent/datasets/字段加工样本.csv/profile")
        assert profile_response.status_code == 200, profile_response.text
        source_sha256 = profile_response.json()["source_sha256"]
        source_path = Path(os.environ["AGENTFLOW_DATA_WORKSPACE_DIR"]) / "字段加工样本.csv"
        source_hash_before = sha256(source_path.read_bytes()).hexdigest()

        base = {"dataset_name": "字段加工样本.csv", "source_sha256": source_sha256}
        cases = [
            {**base, "operation_type": "arithmetic", "primary_column": "单价", "secondary_column": "数量", "result_column": "金额", "arithmetic_operator": "multiply"},
            {**base, "operation_type": "date_part", "primary_column": "日期", "result_column": "月份", "date_part": "month"},
            {**base, "operation_type": "round_number", "primary_column": "单价", "result_column": "单价保留两位", "round_digits": 2},
            {**base, "operation_type": "rank", "primary_column": "成绩", "result_column": "成绩排名"},
            {**base, "operation_type": "share", "primary_column": "单价", "result_column": "单价占比"},
            {**base, "operation_type": "segment", "primary_column": "成绩", "result_column": "成绩分档"},
            {**base, "operation_type": "cumulative", "primary_column": "单价", "secondary_column": "日期", "result_column": "单价累计"},
            {**base, "operation_type": "period_change", "primary_column": "单价", "secondary_column": "日期", "result_column": "单价环比"},
            {**base, "operation_type": "period_rate", "primary_column": "单价", "secondary_column": "日期", "result_column": "单价环比百分比"},
            {**base, "operation_type": "text_trim", "primary_column": "姓名", "result_column": "姓名规范化"},
        ]
        previews = [_preview(client, request) for request in cases]
        assert previews[2]["plan"]["parameters"]["digits"] == 2
        assert previews[5]["plan"]["parameters"]["strategy"] == "score_bands"
        assert previews[7]["empty_result_count"] >= 1
        assert previews[8]["empty_result_count"] >= 1
        assert previews[9]["previews"][0]["result_value"] == "Alice"

        divide = _preview(
            client,
            {**base, "operation_type": "arithmetic", "primary_column": "单价", "secondary_column": "数量", "result_column": "单价除数量", "arithmetic_operator": "divide"},
        )
        assert divide["empty_result_count"] >= 1
        assert any("为 0" in warning for warning in divide["warnings"])
        unknown = client.post(
            "/api/agents/data_agent/transformations/preview",
            json={**base, "operation_type": "rank", "primary_column": "不存在字段"},
        )
        assert unknown.status_code == 400, unknown.text
        no_operation = client.post(
            "/api/agents/data_agent/transformations/preview",
            json={**base, "goal": "帮我计算金额"},
        )
        assert no_operation.status_code == 400, no_operation.text
        assert "字段向导" in no_operation.json()["detail"]

        start = client.post(
            "/api/agents/data_agent/transformations/export/start",
            json={**cases[0], "confirmed": True},
        )
        assert start.status_code == 202, start.text
        result = _wait_for_terminal_result(client, start.json()["task_id"])
        assert result["status"] == "completed", result
        assert result["verification"]["passed"] is True
        assert result["verification"]["result_column"] == "金额"
        assert result["artifact"]["uri"].startswith("agentflow-output://data_transformations/")

        output_path = Path(os.environ["AGENTFLOW_DATA_TRANSFORMATION_OUTPUT_DIR"]) / result["artifact"]["name"]
        # CSV 输入必须维持 CSV 副本：客户拿到的是原表加新字段，而不是被强行包装成多工作表报告。
        assert output_path.suffix == ".csv"
        with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        assert "金额" in (rows[0] if rows else {})
        assert len(rows) == 4
        assert sha256(source_path.read_bytes()).hexdigest() == source_hash_before

        history = client.get(f"/api/tasks/{result['task_id']}/artifacts")
        assert history.status_code == 200, history.text
        artifact = history.json()["artifacts"][0]
        assert artifact["metadata"]["output_path"] == "<hidden>"
        assert artifact["metadata"]["result_column"] == "金额"
        retry = client.post(f"/api/tasks/{result['task_id']}/retry")
        assert retry.status_code == 200 and retry.json()["accepted"] is False

        # 多字段队列必须在同一份副本里追加不同的新列：预览、导出、回读和历史均不能只保留
        # 第一项。队列仍只引用已画像的原字段，避免悄悄演化为任意公式链。
        batch_operations = [
            {
                "operation_type": "arithmetic",
                "primary_column": "单价",
                "secondary_column": "数量",
                "result_column": "金额",
                "arithmetic_operator": "multiply",
            },
            {
                "operation_type": "date_part",
                "primary_column": "日期",
                "result_column": "月份",
                "date_part": "month",
            },
            {
                "operation_type": "rank",
                "primary_column": "成绩",
                "result_column": "成绩排名",
            },
            {
                "operation_type": "round_number",
                "primary_column": "单价",
                "result_column": "单价保留两位",
                "round_digits": 2,
            },
            {
                "operation_type": "period_rate",
                "primary_column": "单价",
                "secondary_column": "日期",
                "result_column": "单价环比百分比",
            },
        ]
        batch_request = {**base, **batch_operations[0], "operations": batch_operations}
        batch_preview = _preview(client, batch_request)
        assert [plan["result_column"] for plan in batch_preview["plans"]] == ["金额", "月份", "成绩排名", "单价保留两位", "单价环比百分比"]
        assert batch_preview["affected_count"] > batch_preview["row_count"]

        batch_start = client.post(
            "/api/agents/data_agent/transformations/export/start",
            json={**batch_request, "confirmed": True},
        )
        assert batch_start.status_code == 202, batch_start.text
        batch_result = _wait_for_terminal_result(client, batch_start.json()["task_id"])
        assert batch_result["status"] == "completed", batch_result
        assert len(batch_result["plans"]) == 5
        assert batch_result["verification"]["result_columns"] == ["金额", "月份", "成绩排名", "单价保留两位", "单价环比百分比"]
        batch_output = Path(os.environ["AGENTFLOW_DATA_TRANSFORMATION_OUTPUT_DIR"]) / batch_result["artifact"]["name"]
        with batch_output.open("r", encoding="utf-8-sig", newline="") as handle:
            batch_rows = list(csv.DictReader(handle))
        assert all(column in (batch_rows[0] if batch_rows else {}) for column in ["金额", "月份", "成绩排名", "单价保留两位", "单价环比百分比"])
        batch_artifact = client.get(f"/api/tasks/{batch_result['task_id']}/artifacts").json()["artifacts"][0]
        assert batch_artifact["metadata"]["result_columns"] == ["金额", "月份", "成绩排名", "单价保留两位", "单价环比百分比"]

        # Excel 输入仍保留 Excel：只有一张无样式数据副本表，所有新增字段横向追加在原字段之后。
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "实验记录"
        sheet.append(["日期", "测量值"])
        sheet.append(["2026-01-01", 12])
        sheet.append(["2026-02-01", 18])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        imported_xlsx = client.post(
            "/api/agents/data_agent/datasets",
            json={"filename": "实验记录.xlsx", "content_base64": _base64(buffer.getvalue())},
        )
        assert imported_xlsx.status_code == 200, imported_xlsx.text
        xlsx_profile = client.get("/api/agents/data_agent/datasets/实验记录.xlsx/profile").json()
        xlsx_start = client.post(
            "/api/agents/data_agent/transformations/export/start",
            json={
                "dataset_name": "实验记录.xlsx",
                "source_sha256": xlsx_profile["source_sha256"],
                "operation_type": "rank",
                "primary_column": "测量值",
                "result_column": "测量排名",
                "confirmed": True,
            },
        )
        assert xlsx_start.status_code == 202, xlsx_start.text
        xlsx_result = _wait_for_terminal_result(client, xlsx_start.json()["task_id"])
        assert xlsx_result["status"] == "completed", xlsx_result
        xlsx_output = Path(os.environ["AGENTFLOW_DATA_TRANSFORMATION_OUTPUT_DIR"]) / xlsx_result["artifact"]["name"]
        assert xlsx_output.suffix == ".xlsx"
        xlsx_copy = load_workbook(xlsx_output, read_only=False, data_only=False)
        try:
            assert xlsx_copy.sheetnames == ["数据副本"]
            copied_sheet = xlsx_copy["数据副本"]
            copied_headers = [cell.value for cell in copied_sheet[1]]
            assert copied_headers == ["日期", "测量值", "测量排名"]
            assert copied_sheet.max_row == 3
            assert not copied_sheet.tables
            assert copied_sheet["A1"].fill.fill_type is None
        finally:
            xlsx_copy.close()

        # 尚未运行的取消可以稳定落库，不创建或登记任何新的数据副本。
        cancel_id = "task_data_transform_abcdef123456"
        cancel_request = DataTransformationExportRequest(**{**cases[1], "confirmed": True})
        create_data_transformation_queued_run(task_id=cancel_id, request=cancel_request)
        cancelled = asyncio.run(cancel_data_transformation_task(cancel_id))
        assert cancelled is not None and cancelled.accepted is True
        cancelled_result = client.get(f"/api/agents/data_agent/transformations/export/{cancel_id}/result")
        assert cancelled_result.status_code == 200
        assert cancelled_result.json()["status"] == "cancelled"
        assert cancelled_result.json()["artifact"] is None

    print(
        "Data transformation verification passed: "
        "ten_operations=true batch_five_fields=true csv_and_plain_xlsx=validated history=hidden-path cancel=safe source_unchanged=true"
    )


if __name__ == "__main__":
    main()
