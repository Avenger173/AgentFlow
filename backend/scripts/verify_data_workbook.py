"""数据工作台 D3 的正式 Excel 交付离线回归。

验证受控导出会新建工作簿、保留原始/清洗/分析表、写入可编辑原生 Table 与 Chart，并且只返回
逻辑 artifact URI。脚本不打印任何单元格内容、绝对路径或真实客户数据。
"""

from __future__ import annotations

import atexit
import base64
from hashlib import sha256
from io import BytesIO
import os
from pathlib import Path
import shutil
import sys
import tempfile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1]
VERIFY_ROOT = Path(tempfile.mkdtemp(prefix="agentflow_data_workbook_verify_"))
VERIFY_DATA_DIR = VERIFY_ROOT / "data"
VERIFY_OUTPUT_DIR = VERIFY_ROOT / "output" / "data_analysis"
os.environ["AGENTFLOW_DATA_DIR"] = str(VERIFY_DATA_DIR)
os.environ["AGENTFLOW_DATA_ANALYSIS_OUTPUT_DIR"] = str(VERIFY_OUTPUT_DIR)
os.environ["AGENTFLOW_CHAT_MODE"] = "mock"
atexit.register(lambda: shutil.rmtree(VERIFY_ROOT, ignore_errors=True))
sys.path.insert(0, str(BACKEND_ROOT))

from main import app  # noqa: E402


def _base64(data: bytes) -> str:
    return base64.b64encode(data).decode("ascii")


def _sample_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售明细"
    sheet.append(["2026 年模拟销售明细"])
    sheet.append(["日期", "区域", "产品", "金额", "数量", "订单号"])
    sheet.append(["2026-01-05", "华东", "A 产品", 1200, 12, "A-001"])
    sheet.append(["2026-01-22", "华南", "A 产品", 800, 8, "A-002"])
    sheet.append(["2026-02-10", "华东", "B 产品", 1500, 15, "A-003"])
    sheet.append(["2026-02-28", "华北", "B 产品", 900, 9, "A-004"])
    sheet.append(["2026-03-08", "华东", "A 产品", 1800, 18, "A-005"])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _assert_native_workbook(path: Path) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        required = {"分析说明", "数据概览", "质量问题", "原始数据", "清洗数据", "图表"}
        assert required.issubset(set(workbook.sheetnames))
        table_count = sum(len(sheet.tables) for sheet in workbook.worksheets)
        chart_count = sum(len(sheet._charts) for sheet in workbook.worksheets)
        assert table_count >= 6
        assert chart_count >= 2
        # “数据概览”第一条指标必须是数值单元格，而不是界面格式化文本。
        assert isinstance(workbook["数据概览"].cell(row=2, column=2).value, (int, float))
        return table_count, chart_count
    finally:
        workbook.close()


def main() -> None:
    source = _sample_xlsx()
    source_path = VERIFY_ROOT / "source.xlsx"
    source_path.write_bytes(source)
    source_hash = sha256(source_path.read_bytes()).hexdigest()

    with TestClient(app) as client:
        imported = client.post(
            "/api/agents/data_agent/datasets",
            json={"filename": "销售交付样本.xlsx", "content_base64": _base64(source)},
        )
        assert imported.status_code == 200, imported.text

        preview = client.post(
            "/api/agents/data_agent/analysis/preview",
            json={
                "dataset_name": "销售交付样本.xlsx",
                "goal": "分析月度销售趋势、区域表现和产品结构",
                "max_chart_count": 3,
            },
        )
        assert preview.status_code == 200, preview.text
        source_sha256 = preview.json()["dataset_profile"]["source_sha256"]

        rejected = client.post(
            "/api/agents/data_agent/analysis/export",
            json={
                "dataset_name": "销售交付样本.xlsx",
                "source_sha256": source_sha256,
                "goal": "分析月度销售趋势、区域表现和产品结构",
                "confirmed": False,
            },
        )
        assert rejected.status_code == 422

        wrong_hash = client.post(
            "/api/agents/data_agent/analysis/export",
            json={
                "dataset_name": "销售交付样本.xlsx",
                "source_sha256": "0" * 64,
                "goal": "分析月度销售趋势、区域表现和产品结构",
                "confirmed": True,
            },
        )
        assert wrong_hash.status_code == 400
        assert not list(VERIFY_OUTPUT_DIR.glob("*.xlsx")) if VERIFY_OUTPUT_DIR.exists() else True

        export = client.post(
            "/api/agents/data_agent/analysis/export",
            json={
                "dataset_name": "销售交付样本.xlsx",
                "source_sha256": source_sha256,
                "goal": "分析月度销售趋势、区域表现和产品结构",
                "max_chart_count": 3,
                "confirmed": True,
            },
        )
        assert export.status_code == 200, export.text
        payload = export.json()
        assert payload["verification"]["passed"] is True
        assert payload["artifact"]["uri"].startswith("agentflow-output://data_analysis/")
        assert str(VERIFY_ROOT) not in str(payload)
        output_path = VERIFY_OUTPUT_DIR / payload["artifact"]["name"]
        assert output_path.exists()
        table_count, chart_count = _assert_native_workbook(output_path)

        second_export = client.post(
            "/api/agents/data_agent/analysis/export",
            json={
                "dataset_name": "销售交付样本.xlsx",
                "source_sha256": source_sha256,
                "goal": "分析月度销售趋势、区域表现和产品结构",
                "max_chart_count": 3,
                "confirmed": True,
            },
        )
        assert second_export.status_code == 200, second_export.text
        assert second_export.json()["artifact"]["name"] != payload["artifact"]["name"]

    assert sha256(source_path.read_bytes()).hexdigest() == source_hash
    print(
        "Data workbook verification passed: "
        f"native_tables={table_count} native_charts={chart_count} source_unchanged=true artifact_uri=true"
    )


if __name__ == "__main__":
    main()
