"""数据工作台 D1 的受控导入、表头识别与确定性画像服务。

这里有意不出现模型、网络、SQL、公式或任意表达式。服务只读取客户端明确导入到固定目录
的单个 Excel/CSV，并把行级数据限制在有限预览响应中；普通日志、任务历史和后续模型上下文
只应保留本模块返回的结构画像和统计摘要。
"""

from __future__ import annotations

import base64
import csv
from collections import Counter, OrderedDict
from datetime import UTC, date, datetime
from hashlib import sha256
from pathlib import Path
import re
from threading import RLock
from typing import Any

from app.core.config import settings
from app.schemas.data_agent import (
    DataColumnProfile,
    DataDatasetInfo,
    DataDatasetProfileResponse,
    DataHeaderCandidate,
    DataQualitySummary,
    DataSheetInfo,
)

try:  # requirements 缺失时返回可操作的错误，而不是导入接口 500。
    import pandas as pd
except ImportError:  # pragma: no cover - 正常安装 requirements 时不会走到这里。
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - 正常安装 requirements 时不会走到这里。
    load_workbook = None


DATASET_SUFFIXES = {".xlsx", ".csv"}
MAX_DATASET_BYTES = 20_000_000
MAX_DATASET_ROWS = 100_000
MAX_DATASET_COLUMNS = 100
MAX_VISIBLE_SHEETS = 10
PREVIEW_ROW_LIMIT = 20
PREVIEW_COLUMN_LIMIT = 20
HEADER_SCAN_ROWS = 10
PROFILE_CACHE_MAX_ENTRIES = 8


class DataWorkspaceError(ValueError):
    """面向客户的受控数据工作区错误。"""


def data_workspace_dependency_status() -> dict[str, object]:
    """返回数据工作台解析能力，供启动健康检查提前提示环境问题。

    这个检查只复用模块导入结果，不扫描文件、不创建 DataFrame，也不读取客户导入的资料。
    因此可以安全放在轻量 `/health` 路径中，避免客户导入成功后才发现无法建立画像。
    """

    missing: list[str] = []
    if pd is None:
        missing.append("pandas")
    if load_workbook is None:
        missing.append("openpyxl")

    if not missing:
        return {
            "ready": True,
            "message": "数据工作台解析依赖已就绪。",
        }

    missing_text = "、".join(missing)
    return {
        "ready": False,
        "message": (
            f"数据工作台依赖未安装：{missing_text}。"
            "请在 backend 目录使用桌面端实际 Python 执行 pip install -r requirements.txt。"
        ),
    }


# 文件版本键由绝对路径、mtime_ns、size 构成。导入文件是不可覆盖的新副本，且用户可以主动
# 刷新列表；因此缓存不会跨版本返回旧画像，也避免每次切换 Qt 页面都重复解析 20MB 表格。
_profile_cache: OrderedDict[tuple[str, int, int], DataDatasetProfileResponse] = OrderedDict()
_profile_cache_lock = RLock()


def data_workspace_dir() -> Path:
    """返回内部受控数据目录，绝对路径永不通过 API 返回给客户端。"""

    return settings.data_workspace_dir


def list_data_datasets() -> list[DataDatasetInfo]:
    """列出客户显式导入的数据文件；不在列表请求里解析内容。"""

    root = data_workspace_dir()
    if not root.exists():
        return []
    return [
        _dataset_info(path)
        for path in sorted(root.iterdir(), key=lambda item: item.name.casefold())
        if path.is_file() and path.suffix.lower() in DATASET_SUFFIXES
    ]


def import_data_dataset_base64(*, filename: str, content_base64: str) -> DataDatasetInfo:
    """解码并写入受控目录；同名文件自动创建新副本，绝不覆盖源数据。"""

    safe_name = _safe_dataset_filename(filename)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in DATASET_SUFFIXES:
        raise DataWorkspaceError("数据工作台 D1 只支持 .xlsx 或 .csv 文件。")
    try:
        data = base64.b64decode(content_base64.encode("ascii"), validate=True)
    except (UnicodeEncodeError, ValueError) as exc:
        raise DataWorkspaceError("数据文件内容不是有效的 Base64 数据。") from exc
    if not data:
        raise DataWorkspaceError("数据文件为空，无法建立数据画像。")
    if len(data) > MAX_DATASET_BYTES:
        raise DataWorkspaceError("单个数据文件最大支持 20MB，请先拆分或筛选后再导入。")

    root = data_workspace_dir()
    root.mkdir(parents=True, exist_ok=True)
    target = _next_available_path(root, safe_name)
    try:
        target.write_bytes(data)
    except OSError as exc:
        raise DataWorkspaceError("无法写入受控数据工作区，请检查目录权限和可用空间。") from exc
    return _dataset_info(target)


def get_data_dataset_profile(dataset_name: str) -> DataDatasetProfileResponse:
    """读取一份已导入数据集的确定性画像，并返回受限的本地预览。"""

    _ensure_parser_dependencies()
    path = _data_dataset_path(dataset_name)
    stat = path.stat()
    if stat.st_size > MAX_DATASET_BYTES:
        raise DataWorkspaceError("数据文件超过 20MB 限制，不能建立画像。")
    cache_key = (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
    with _profile_cache_lock:
        cached = _profile_cache.get(cache_key)
        if cached is not None:
            _profile_cache.move_to_end(cache_key)
            return cached.model_copy(deep=True)

    profile = _build_dataset_profile(path)
    with _profile_cache_lock:
        _profile_cache[cache_key] = profile
        _profile_cache.move_to_end(cache_key)
        while len(_profile_cache) > PROFILE_CACHE_MAX_ENTRIES:
            _profile_cache.popitem(last=False)
    return profile.model_copy(deep=True)


def load_data_dataset_for_analysis(dataset_name: str) -> tuple[DataDatasetProfileResponse, Any]:
    """为 D2/D3 的本地 Tool 读取受控表格。

    这个函数刻意不通过 FastAPI 暴露，也不向模型返回 ``frame``。D1 画像先决定推荐工作表、
    表头和稳定列名，D2 只能在相同的受控副本上做白名单计算，避免两套解析规则得出不同事实。
    """

    _ensure_parser_dependencies()
    profile = get_data_dataset_profile(dataset_name)
    path = _data_dataset_path(dataset_name)
    try:
        if path.suffix.lower() == ".xlsx":
            frame = pd.read_excel(
                path,
                sheet_name=profile.selected_sheet,
                header=profile.header_row - 1,
                dtype=object,
                nrows=MAX_DATASET_ROWS + 1,
                engine="openpyxl",
            )
        else:
            encoding, delimiter = _detect_csv_transport(path)
            frame = pd.read_csv(
                path,
                encoding=encoding,
                sep=delimiter,
                dtype=object,
                nrows=MAX_DATASET_ROWS + 1,
                keep_default_na=True,
                na_values=[""],
            )
    except Exception as exc:
        raise DataWorkspaceError("无法再次读取已导入的数据文件，请重新导入后再分析。") from exc

    _validate_frame_shape(frame)
    stable_columns = [column.name for column in profile.columns]
    if frame.shape[1] != len(stable_columns):
        # 受控副本原则上不可覆盖。若磁盘内容被外部异常修改，拒绝继续而不是把错列交给计算器。
        raise DataWorkspaceError("数据文件结构与已建立的画像不一致，请重新导入后再分析。")
    frame.columns = stable_columns
    return profile, frame.copy(deep=True)


def _build_dataset_profile(path: Path) -> DataDatasetProfileResponse:
    """按格式读取最多 100001 行，明确拒绝超限而不静默截断。"""

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        sheets, selected_sheet, header_row = _inspect_xlsx(path)
        try:
            frame = pd.read_excel(
                path,
                sheet_name=selected_sheet,
                header=header_row - 1,
                dtype=object,
                nrows=MAX_DATASET_ROWS + 1,
                engine="openpyxl",
            )
        except Exception as exc:  # openpyxl 的加密/损坏文件异常类型不稳定，统一为可理解错误。
            raise DataWorkspaceError("无法读取 Excel 文件；请确认它未加密、未损坏且包含有效数据表。") from exc
    else:
        encoding, delimiter = _detect_csv_transport(path)
        selected_sheet = "CSV 数据"
        header_row = 1
        sheets = [
            DataSheetInfo(
                name=selected_sheet,
                row_count=0,
                column_count=0,
                recommended=True,
                header_candidates=[DataHeaderCandidate(row_number=1, score=1.0, non_empty_cells=0)],
            )
        ]
        try:
            frame = pd.read_csv(
                path,
                encoding=encoding,
                sep=delimiter,
                dtype=object,
                nrows=MAX_DATASET_ROWS + 1,
                keep_default_na=True,
                na_values=[""],
            )
        except Exception as exc:
            raise DataWorkspaceError("无法按 UTF-8、GB18030 或 UTF-16 和常见分隔符读取 CSV，请检查编码与分隔符。") from exc

    _validate_frame_shape(frame)

    # pandas 会为重复列名追加 .1；D1 同时保留这一安全可用名称，并记录重复表头提醒。
    # 必须在稳定命名之前计数，否则同名字段会被 pandas 的后缀悄悄掩盖。
    raw_column_names = [_safe_column_name(value, index) for index, value in enumerate(frame.columns, start=1)]
    duplicate_header_count = _duplicate_header_count(raw_column_names)
    column_names = list(raw_column_names)
    frame.columns = column_names
    column_profiles = [_profile_column(index, name, frame.iloc[:, index - 1]) for index, name in enumerate(column_names, start=1)]
    missing_cells = sum(item.missing_count for item in column_profiles)
    empty_columns = sum(1 for item in column_profiles if item.non_null_count == 0)
    parse_issue_columns = sum(1 for item in column_profiles if item.parse_issue_count > 0)
    # 只在 100k x 100 的硬边界内做一次标准化后去重；不把标准化结果持久化或回传。
    # 逐列转成稳定文本后再判重，避免 pandas 对 object 列 ``fillna`` 的下推行为在将来版本变化。
    normalized = frame.map(_normalise_cell_for_duplicate_check)
    duplicate_rows = int(normalized.duplicated().sum())

    warnings: list[str] = []
    if duplicate_header_count:
        warnings.append(f"发现 {duplicate_header_count} 个重复表头；后续分析会使用稳定列名，请在 D2 确认映射。")
    if empty_columns:
        warnings.append(f"发现 {empty_columns} 个空列；D2 默认不会把它们纳入指标计算。")
    if duplicate_rows:
        warnings.append(f"发现 {duplicate_rows} 行完全重复数据；D1 仅提示，不会自动删除。")
    if parse_issue_columns:
        warnings.append(f"有 {parse_issue_columns} 列存在疑似数值或日期格式不一致；D2 需要先确认安全清洗方式。")

    if suffix == ".csv":
        sheets[0] = sheets[0].model_copy(
            update={
                "row_count": int(frame.shape[0]) + header_row,
                "column_count": int(frame.shape[1]),
                "header_candidates": [
                    DataHeaderCandidate(
                        row_number=1,
                        score=1.0,
                        non_empty_cells=int(frame.shape[1]),
                        preview=column_names[:6],
                    )
                ],
            }
        )

    preview_columns = column_names[:PREVIEW_COLUMN_LIMIT]
    preview_rows = [
        [_display_cell(value) for value in row]
        for row in frame.iloc[:PREVIEW_ROW_LIMIT, :PREVIEW_COLUMN_LIMIT].itertuples(index=False, name=None)
    ]
    return DataDatasetProfileResponse(
        dataset=_dataset_info(path),
        source_sha256=_sha256_file(path),
        selected_sheet=selected_sheet,
        header_row=header_row,
        row_count=int(frame.shape[0]),
        column_count=int(frame.shape[1]),
        sheets=sheets,
        columns=column_profiles,
        preview_columns=preview_columns,
        preview_rows=preview_rows,
        quality_summary=DataQualitySummary(
            missing_cell_count=missing_cells,
            duplicate_row_count=duplicate_rows,
            empty_column_count=empty_columns,
            duplicate_header_count=duplicate_header_count,
            parse_issue_column_count=parse_issue_columns,
        ),
        warnings=warnings,
    )


def _validate_frame_shape(frame: Any) -> None:
    """统一 D1 画像和 D2 本地读取的硬边界，禁止任一入口静默截断。"""

    if frame.shape[0] > MAX_DATASET_ROWS:
        raise DataWorkspaceError("数据行超过 100,000 行限制；请先筛选或拆分后再导入。")
    if frame.shape[1] > MAX_DATASET_COLUMNS:
        raise DataWorkspaceError("数据列超过 100 列限制；请先保留本次需要的字段后再导入。")
    if frame.shape[1] == 0:
        raise DataWorkspaceError("未识别到数据列，请确认首行是稳定表头。")
    if frame.shape[0] == 0:
        raise DataWorkspaceError("已识别表头，但没有数据行，暂时无法建立画像。")


def _inspect_xlsx(path: Path) -> tuple[list[DataSheetInfo], str, int]:
    """只扫描可见工作表和前十行，选择最像主数据表的候选，不读取整本工作簿。"""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise DataWorkspaceError("无法打开 Excel 文件；加密、损坏或不受支持的工作簿不能导入。") from exc
    try:
        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if not visible:
            raise DataWorkspaceError("Excel 中没有可见工作表，无法选择主数据表。")
        if len(visible) > MAX_VISIBLE_SHEETS:
            raise DataWorkspaceError("Excel 可见工作表超过 10 个；请拆分后再导入。")
        inspected: list[tuple[DataSheetInfo, int]] = []
        for sheet in visible:
            max_rows = max(0, int(sheet.max_row or 0))
            max_columns = max(0, int(sheet.max_column or 0))
            candidates = _header_candidates(sheet, max_columns)
            # max_row/max_column 是 Excel 的占用范围，可能含格式尾部；仅用于主表推荐，真正边界
            # 仍由 pandas 实读数据行裁决。
            score = max_rows * max_columns if max_rows >= 2 and max_columns >= 1 else -1
            inspected.append(
                (
                    DataSheetInfo(
                        name=sheet.title,
                        row_count=max_rows,
                        column_count=max_columns,
                        header_candidates=candidates,
                    ),
                    score,
                )
            )
        recommended_index = max(range(len(inspected)), key=lambda index: inspected[index][1])
        sheets = [
            info.model_copy(update={"recommended": index == recommended_index})
            for index, (info, _score) in enumerate(inspected)
        ]
        recommended = sheets[recommended_index]
        if recommended.row_count < 2 or recommended.column_count < 1:
            raise DataWorkspaceError("Excel 没有包含表头和数据行的可用工作表。")
        header_row = recommended.header_candidates[0].row_number if recommended.header_candidates else 1
        return sheets, recommended.name, header_row
    finally:
        workbook.close()


def _header_candidates(sheet: Any, max_columns: int) -> list[DataHeaderCandidate]:
    """用非空、文本和唯一性给前十行打分，避免把模型判断带入 D1。"""

    candidates: list[DataHeaderCandidate] = []
    width = min(max_columns, MAX_DATASET_COLUMNS)
    for row_number in range(1, min(int(sheet.max_row or 0), HEADER_SCAN_ROWS) + 1):
        values = [sheet.cell(row=row_number, column=index).value for index in range(1, width + 1)]
        visible = [_display_cell(value) for value in values if not _is_missing(value)]
        if not visible:
            continue
        text_count = sum(1 for value in values if isinstance(value, str) and value.strip())
        unique_count = len({value.casefold() for value in visible})
        score = round((len(visible) + text_count * 0.8 + unique_count * 0.4) / max(1, width), 3)
        candidates.append(
            DataHeaderCandidate(
                row_number=row_number,
                score=max(score, 0.0),
                non_empty_cells=len(visible),
                preview=visible[:6],
            )
        )
    candidates.sort(key=lambda item: (-item.score, item.row_number))
    return candidates[:3]


def _detect_csv_transport(path: Path) -> tuple[str, str]:
    """在常见办公 CSV 编码和分隔符中做有限识别。"""

    raw = path.read_bytes()[:65_536]
    text = ""
    selected_encoding = ""
    # Excel/WPS 等工具可能把“CSV”导出为 UTF-16 文本。UTF-16 的低位 NUL 对 UTF-8 来说
    # 也是合法字节，若先盲试 UTF-8 就会把表头和分隔符识别成乱码；因此先看 BOM 或稳定的
    # 双字节 NUL 分布，再回落到 UTF-8/GB18030。
    utf16_encoding = _detect_utf16_encoding(raw)
    candidates = ([utf16_encoding] if utf16_encoding else []) + ["utf-8-sig", "utf-8", "gb18030"]
    for encoding in candidates:
        sample = raw
        # 采样长度可能刚好落在 UTF-16 code unit 的中间；只为嗅探去掉最后一个残缺字节，
        # 实际 pandas 读取仍使用完整原文件，不会丢数据。
        if encoding.startswith("utf-16") and len(sample) % 2:
            sample = sample[:-1]

        # 64KB 采样同样可能从 UTF-8/GB18030 的多字节字符中间截断。这个情形不代表原文件
        # 编码坏了，只能在样本末尾有限回退；绝不使用 errors="ignore" 掩盖样本内部的非法字节。
        max_trailing_bytes = 3 if encoding in {"utf-8-sig", "utf-8", "gb18030"} else 0
        for trailing_bytes in range(max_trailing_bytes + 1):
            candidate = sample if trailing_bytes == 0 else sample[:-trailing_bytes]
            try:
                text = candidate.decode(encoding)
                selected_encoding = encoding
                break
            except UnicodeDecodeError:
                continue
        if selected_encoding:
            break
    if not selected_encoding:
        raise DataWorkspaceError("CSV 编码无法识别；请另存为 UTF-8、GB18030 或 UTF-16 后再导入。")
    try:
        dialect = csv.Sniffer().sniff(text, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        first_line = next((line for line in text.splitlines() if line.strip()), "")
        counts = {candidate: first_line.count(candidate) for candidate in (",", "\t", ";", "|")}
        delimiter = max(counts, key=counts.get)
        # 单列 CSV 没有任何分隔符也应能作为数据集导入；逗号是其无歧义的默认解析器参数。
        if counts[delimiter] == 0:
            delimiter = ","
    return selected_encoding, delimiter


def _detect_utf16_encoding(raw: bytes) -> str:
    """识别带 BOM 或 ASCII 表头主导的 UTF-16 CSV，不把任意二进制误判为文本。"""

    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        # ``utf-16`` 会依据 BOM 自动选择字节序，并在解析时忽略 BOM 本身。
        return "utf-16"
    if len(raw) < 8:
        return ""

    sample = raw[: min(len(raw), 4_096)]
    even_bytes = sample[::2]
    odd_bytes = sample[1::2]
    even_null_ratio = even_bytes.count(0) / max(1, len(even_bytes))
    odd_null_ratio = odd_bytes.count(0) / max(1, len(odd_bytes))
    # 无 BOM 的 UTF-16 CSV 通常仍含有 ASCII 标题、分隔符和换行；要求一侧 NUL 占比显著
    # 高于另一侧，避免把带偶然 0x00 的损坏/二进制文件误当可分析表格。
    if odd_null_ratio >= 0.30 and even_null_ratio <= 0.05:
        return "utf-16-le"
    if even_null_ratio >= 0.30 and odd_null_ratio <= 0.05:
        return "utf-16-be"
    return ""


def _profile_column(index: int, name: str, series: Any) -> DataColumnProfile:
    values = list(series.tolist())
    present = [value for value in values if not _is_missing(value)]
    display_values = [_display_cell(value) for value in present]
    non_null_count = len(present)
    missing_count = len(values) - non_null_count
    unique_count = len(set(display_values))
    if not present:
        return DataColumnProfile(
            index=index,
            name=name,
            inferred_type="text",
            non_null_count=0,
            missing_count=missing_count,
            unique_count=0,
            parse_issue_count=0,
        )

    if all(isinstance(value, bool) for value in present):
        return DataColumnProfile(
            index=index,
            name=name,
            inferred_type="boolean",
            non_null_count=non_null_count,
            missing_count=missing_count,
            unique_count=unique_count,
            parse_issue_count=0,
        )

    numeric_values = [_parse_number(value) for value in present]
    numeric_count = sum(value is not None for value in numeric_values)
    numeric_ratio = numeric_count / non_null_count
    if numeric_ratio >= 0.92:
        accepted = [value for value in numeric_values if value is not None]
        return DataColumnProfile(
            index=index,
            name=name,
            inferred_type="number",
            non_null_count=non_null_count,
            missing_count=missing_count,
            unique_count=unique_count,
            parse_issue_count=non_null_count - numeric_count,
            numeric_min=round(min(accepted), 6),
            numeric_max=round(max(accepted), 6),
            numeric_mean=round(sum(accepted) / len(accepted), 6),
        )

    parsed_dates = [_parse_date(value) for value in present]
    date_count = sum(value is not None for value in parsed_dates)
    date_ratio = date_count / non_null_count
    date_signal = any(_looks_like_date(value) for value in present) or bool(
        re.search(r"date|time|日期|时间|月份|季度|年份", name, re.IGNORECASE)
    )
    if date_signal and date_ratio >= 0.9:
        accepted_dates = [value for value in parsed_dates if value is not None]
        return DataColumnProfile(
            index=index,
            name=name,
            inferred_type="date",
            non_null_count=non_null_count,
            missing_count=missing_count,
            unique_count=unique_count,
            parse_issue_count=non_null_count - date_count,
            earliest=min(accepted_dates).isoformat(),
            latest=max(accepted_dates).isoformat(),
        )

    inferred_type = "mixed" if (numeric_count > 0 or (date_signal and date_count > 0)) else "text"
    parse_issues = max(numeric_count, date_count) if inferred_type == "mixed" else 0
    return DataColumnProfile(
        index=index,
        name=name,
        inferred_type=inferred_type,
        non_null_count=non_null_count,
        missing_count=missing_count,
        unique_count=unique_count,
        parse_issue_count=parse_issues,
    )


def _parse_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    text = _display_cell(value).replace(",", "").replace(" ", "")
    # D1 只用于类型识别与范围显示，百分号不擅自改成小数，货币符号仅剥离常见显示字符。
    text = text.replace("¥", "").replace("￥", "").replace("$", "")
    if text.endswith("%"):
        text = text[:-1]
    if not text or not re.fullmatch(r"[-+]?(?:\d+\.?\d*|\.\d+)(?:[eE][-+]?\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _display_cell(value)
    if not _looks_like_date(text):
        return None
    try:
        parsed = pd.to_datetime(text, errors="coerce")
    except (TypeError, ValueError):
        return None
    if pd.isna(parsed):
        return None
    return parsed.date()


def _looks_like_date(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    text = _display_cell(value)
    return bool(re.search(r"\d{4}[-/.年]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", text))


def _dataset_info(path: Path) -> DataDatasetInfo:
    stat = path.stat()
    return DataDatasetInfo(
        name=path.name,
        relative_path=path.name,
        size_bytes=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime, tz=UTC).isoformat(),
        dataset_type="xlsx" if path.suffix.lower() == ".xlsx" else "csv",
    )


def _safe_dataset_filename(filename: str) -> str:
    raw = filename.strip()
    candidate = Path(raw).name
    if not raw or candidate != raw or candidate in {".", ".."}:
        raise DataWorkspaceError("文件名无效；请重新选择本地 Excel 或 CSV 文件。")
    if not Path(candidate).stem:
        raise DataWorkspaceError("文件名缺少有效名称。")
    return candidate


def _data_dataset_path(dataset_name: str) -> Path:
    safe_name = _safe_dataset_filename(dataset_name)
    path = (data_workspace_dir() / safe_name).resolve()
    try:
        path.relative_to(data_workspace_dir().resolve())
    except ValueError as exc:
        raise DataWorkspaceError("数据文件路径不在受控工作区内。") from exc
    if path.suffix.lower() not in DATASET_SUFFIXES:
        raise DataWorkspaceError("当前操作只支持 .xlsx 或 .csv 文件。")
    if not path.is_file():
        raise DataWorkspaceError("未找到指定的数据文件；请刷新列表后重试。")
    return path


def _next_available_path(root: Path, filename: str) -> Path:
    candidate = root / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for ordinal in range(2, 10_000):
        alternative = root / f"{stem} ({ordinal}){suffix}"
        if not alternative.exists():
            return alternative
    raise DataWorkspaceError("同名数据文件过多，无法创建新的受控副本。")


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as source:
        while chunk := source.read(1_048_576):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_column_name(value: Any, index: int) -> str:
    name = _display_cell(value).strip()
    return name if name else f"未命名列_{index}"


def _duplicate_header_count(column_names: list[str]) -> int:
    """识别 pandas 为重复表头追加的 ``.1``、``.2`` 后缀。

    D1 只把它当作质量提醒：真实列名仍保留 pandas 已稳定化后的形式，避免后续 Tool 因重名
    无法定位字段。极少数原始字段本来以数字后缀命名时，宁可多提示一次，也不隐瞒歧义。
    """

    normalized: list[str] = []
    for name in column_names:
        match = re.fullmatch(r"(.+)\.(\d+)", name)
        normalized.append(match.group(1) if match else name)
    return sum(count - 1 for count in Counter(normalized).values() if count > 1)


def _display_cell(value: Any) -> str:
    if _is_missing(value):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\x00", "").strip()[:240]


def _normalise_cell_for_duplicate_check(value: Any) -> str:
    return _display_cell(value).casefold()


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _ensure_parser_dependencies() -> None:
    dependency_status = data_workspace_dependency_status()
    if not dependency_status["ready"]:
        raise DataWorkspaceError(str(dependency_status["message"]))
