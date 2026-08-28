"""数据工作台 D3 的 Excel 交付与回读验证服务。

本模块只消费 D2 已通过白名单的确定性计算结果。它不执行客户表达式、不调用模型或网络，
也不修改导入工作区中的源文件；每次导出都会先写入临时工作簿、重新打开验证，再原子地
移动到固定的 ``output/data_analysis`` 目录。
"""

from __future__ import annotations

import math
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.core.config import settings
from app.schemas.data_agent import (
    DataAnalysisPreviewRequest,
    DataAnalysisTable,
    DataWorkbookArtifact,
    DataWorkbookExportRequest,
    DataWorkbookExportResponse,
    DataWorkbookVerification,
)
from app.services.data_analysis import DataAnalysisError, DataAnalysisComputation, compute_data_analysis

_DEPENDENCIES_AVAILABLE = True
try:  # 与 D1/D2 一样：缺少可选依赖时返回可行动错误，而不是 500。
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover - requirements.txt 正常安装时不会进入。
    pd = None
    _DEPENDENCIES_AVAILABLE = False


class DataWorkbookError(ValueError):
    """数据工作簿无法安全生成或验证时抛出的客户可理解错误。"""


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1D4ED8") if _DEPENDENCIES_AVAILABLE else None
_HEADER_FONT = Font(color="FFFFFF", bold=True) if _DEPENDENCIES_AVAILABLE else None
_TABLE_STYLE = "TableStyleMedium2"


def export_data_analysis_workbook(request: DataWorkbookExportRequest) -> DataWorkbookExportResponse:
    """基于当前受控数据创建一个新的 Excel 工作簿，并在返回前回读验证。

    ``confirmed`` 属于 API 契约的一部分，因此这里不再重复确认。路径、文件名和工作簿内容
    均由后端决定；失败时不会留下一份看似成功但未验证的正式交付物。
    """

    _ensure_dependencies()
    try:
        computation = compute_data_analysis(
            DataAnalysisPreviewRequest(
                dataset_name=request.dataset_name,
                goal=request.goal,
                cleaning_policy=request.cleaning_policy,
                max_chart_count=request.max_chart_count,
            )
        )
    except DataAnalysisError as exc:
        raise DataWorkbookError(str(exc)) from exc

    # 上面的导出请求比 D2 请求多了哈希和确认字段，显式转换可保证没有客户端新增字段进入
    # 计算器；禁止把 ``dict`` 直接透传给内部 Tool。
    preview = computation.preview
    if request.source_sha256.casefold() != preview.dataset_profile.source_sha256.casefold():
        raise DataWorkbookError("数据文件已变化或不属于当前预览，请重新生成分析预览后再导出。")

    output_dir = settings.data_analysis_output_dir
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise DataWorkbookError("无法创建数据分析输出目录，请检查磁盘权限和可用空间。") from exc

    destination = _next_output_path(output_dir, request.dataset_name)
    temporary = output_dir / f".{destination.stem}.{uuid.uuid4().hex}.partial.xlsx"
    created_at = datetime.now(timezone.utc).isoformat()
    try:
        rendered = _render_workbook(temporary, computation, request, created_at)
        verification = _verify_workbook(
            temporary,
            expected_sheets=rendered["expected_sheets"],
            minimum_table_count=rendered["minimum_table_count"],
            expected_chart_count=rendered["chart_count"],
            expected_metric_count=len(preview.metrics),
            expected_metric_values=[metric.value for metric in preview.metrics],
        )
        # os.rename 在 Windows 不会覆盖已存在的同名文件；即使极少数并发导出撞名，也宁愿让
        # 本次明确失败，也不冒覆盖他人交付物的风险。
        os.rename(temporary, destination)
    except DataWorkbookError:
        raise
    except FileExistsError as exc:
        raise DataWorkbookError("输出文件命名冲突，请重新确认导出。") from exc
    except OSError as exc:
        raise DataWorkbookError("无法写入或完成数据工作簿，请检查输出目录权限和可用空间。") from exc
    except Exception as exc:
        # 对外仍只给通用安全提示，但保留异常链供本地离线回归和受控日志定位，避免以放宽回读
        # 条件的方式掩盖真实导出错误。
        raise DataWorkbookError("数据工作簿生成或回读验证失败，未生成正式交付文件。") from exc
    finally:
        # 任何失败分支只残留临时文件名；成功后它已被 rename，因此 unlink 是无害的。
        try:
            temporary.unlink(missing_ok=True)
        except OSError:
            pass

    stat = destination.stat()
    return DataWorkbookExportResponse(
        artifact=DataWorkbookArtifact(
            name=destination.name,
            uri=f"agentflow-output://data_analysis/{destination.name}",
            size_bytes=stat.st_size,
            created_at=created_at,
        ),
        verification=verification,
        warnings=preview.warnings,
        skipped_items=preview.skipped_items,
    )


def _ensure_dependencies() -> None:
    if not _DEPENDENCIES_AVAILABLE:
        raise DataWorkbookError("Excel 导出依赖未安装，请在 backend 目录安装 requirements.txt。")


def _render_workbook(
    path: Path,
    computation: DataAnalysisComputation,
    request: DataWorkbookExportRequest,
    created_at: str,
) -> dict[str, Any]:
    """将一次计算结果写为多个职责明确的工作表，返回供回读验证的期望合同。"""

    preview = computation.preview
    workbook = Workbook()
    workbook.remove(workbook.active)
    table_index = 1
    expected_sheets: list[str] = []

    def create_sheet(title: str) -> Any:
        safe_title = _safe_sheet_title(title, {sheet.title for sheet in workbook.worksheets})
        expected_sheets.append(safe_title)
        return workbook.create_sheet(safe_title)

    def add_native_table(sheet: Any) -> None:
        nonlocal table_index
        if sheet.max_column < 1 or sheet.max_row < 2:
            return
        ref = f"A1:{_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName=f"tbl_{table_index:02d}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        table_index += 1

    notes = create_sheet("分析说明")
    notes_rows = [
        ["项目", "内容"],
        ["交付类型", "受控本地数据分析工作簿"],
        ["数据集", preview.dataset_profile.dataset.name],
        ["数据哈希", preview.dataset_profile.source_sha256],
        ["分析目标", request.goal.strip() or "未填写目标，使用结构画像的标准视图。"],
        ["清洗策略", "安全副本：仅去除文本首尾空白，不删除、不填补、不裁剪。"],
        ["生成时间", created_at],
        ["事实边界", "全部指标和图表来自本机确定性计算；未调用模型、联网或客户公式。"],
    ]
    _append_rows(notes, notes_rows)
    _style_sheet(notes, freeze="A2")
    add_native_table(notes)

    overview = create_sheet("数据概览")
    overview_rows = [["指标", "数值", "单位", "汇总方式", "来源字段"]]
    overview_rows.extend(
        [metric.name, _finite_number(metric.value), metric.unit, metric.aggregation, "、".join(metric.source_columns)]
        for metric in preview.metrics
    )
    if len(overview_rows) == 1:
        overview_rows.append(["没有可用指标", "", "", "", ""])
    _append_rows(overview, overview_rows)
    _style_sheet(overview, freeze="A2")
    add_native_table(overview)

    quality = create_sheet("质量问题")
    quality_rows = [["级别", "问题", "影响", "影响数量", "处理方式"]]
    quality_rows.extend(
        [finding.severity, finding.title, finding.impact, finding.affected_count, finding.handling]
        for finding in preview.quality_findings
    )
    if len(quality_rows) == 1:
        quality_rows.append(["info", "未发现可记录的问题", "仍需按业务口径复核。", 0, "未修改源数据。"])
    _append_rows(quality, quality_rows)
    _style_sheet(quality, freeze="A2")
    add_native_table(quality)

    raw = create_sheet("原始数据")
    _write_dataframe(raw, computation.source_frame)
    _style_sheet(raw, freeze="A2")
    add_native_table(raw)

    cleaned = create_sheet("清洗数据")
    _write_dataframe(cleaned, computation.cleaned_frame)
    _style_sheet(cleaned, freeze="A2")
    add_native_table(cleaned)

    analysis_sources: dict[str, tuple[Any, int]] = {}
    for analysis_table in preview.analysis_tables:
        frame = computation.table_frames.get(analysis_table.table_id)
        if frame is None:
            # 不会因为一个意外缺少的分析表生成空白工作表；前面 D2 已记录可行动的跳过原因。
            continue
        sheet = create_sheet(f"分析_{analysis_table.title}")
        _write_dataframe(sheet, frame)
        _style_sheet(sheet, freeze="A2")
        add_native_table(sheet)
        chart_rows = min(len(analysis_table.rows), max(sheet.max_row - 1, 0))
        analysis_sources[analysis_table.table_id] = (sheet, chart_rows)

    chart_count = 0
    if preview.charts:
        charts = create_sheet("图表")
        charts.sheet_view.showGridLines = False
        chart_row = 1
        for contract in preview.charts:
            source = analysis_sources.get(contract.table_id)
            if source is None:
                continue
            source_sheet, chart_rows = source
            if chart_rows < 2:
                continue
            chart = _build_chart(contract.chart_type, contract.title, source_sheet, chart_rows)
            charts.add_chart(chart, f"A{chart_row}")
            chart_row += 16
            chart_count += 1
        if chart_count == 0:
            workbook.remove(charts)
            expected_sheets.remove(charts.title)

    workbook.save(path)
    workbook.close()
    return {
        "expected_sheets": expected_sheets,
        "minimum_table_count": max(table_index - 1, 0),
        "chart_count": chart_count,
    }


def _write_dataframe(sheet: Any, frame: Any) -> None:
    """按原始 Python 数值写入数据帧，而不是导出 UI 格式化字符串。"""

    headers = [str(column) for column in frame.columns]
    sheet.append(headers)
    if frame.empty:
        # D1 允许合法的“只有表头”文件。工作簿不能拥有视觉上的空表，因此保留一条非数据说明，
        # 同时不建立 native Table，避免把说明误认为客户数据。
        sheet.append(["没有可导出的数据记录"] + [None] * max(0, len(headers) - 1))
        return
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_excel_cell(value) for value in row])


def _append_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append([_excel_cell(value) for value in row])


def _style_sheet(sheet: Any, *, freeze: str) -> None:
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 24
    # 只扫描前 80 行，避免为十万行数据的列宽计算制造新的性能瓶颈。
    for column_index in range(1, sheet.max_column + 1):
        longest = 8
        for row_index in range(1, min(sheet.max_row, 80) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            longest = max(longest, min(len(str(value or "")), 42))
        sheet.column_dimensions[_column_letter(column_index)].width = min(max(longest + 2, 11), 42)


def _build_chart(chart_type: str, title: str, source_sheet: Any, row_count: int) -> Any:
    if chart_type == "line":
        chart = LineChart()
        chart.style = 13
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "doughnut":
        chart = DoughnutChart()
        chart.holeSize = 55
    else:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
    chart.title = title
    chart.height = 7.2
    chart.width = 14.0
    # 饼图/环形图没有坐标轴；只为折线和柱状图写入轴标题，避免不同 OpenPyXL 原生对象
    # 接口不一致导致整个工作簿导出失败。
    if chart_type in {"line", "bar"}:
        chart.y_axis.title = source_sheet.cell(row=1, column=2).value
        chart.x_axis.title = source_sheet.cell(row=1, column=1).value
    data = Reference(source_sheet, min_col=2, min_row=1, max_row=row_count + 1)
    categories = Reference(source_sheet, min_col=1, min_row=2, max_row=row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    return chart


def _verify_workbook(
    path: Path,
    *,
    expected_sheets: list[str],
    minimum_table_count: int,
    expected_chart_count: int,
    expected_metric_count: int,
    expected_metric_values: list[float],
) -> DataWorkbookVerification:
    """重新打开临时文件，核验原生对象和写入值后才允许它成为正式 artifact。"""

    try:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    except Exception as exc:
        raise DataWorkbookError("导出的 Excel 无法重新打开，已拒绝登记该文件。") from exc
    try:
        names = workbook.sheetnames
        missing = [name for name in expected_sheets if name not in names]
        if missing:
            raise DataWorkbookError("导出的 Excel 缺少必要工作表，已拒绝登记该文件。")
        table_count = sum(len(sheet.tables) for sheet in workbook.worksheets)
        chart_count = sum(len(sheet._charts) for sheet in workbook.worksheets)
        if table_count < minimum_table_count:
            raise DataWorkbookError("导出的 Excel 缺少原生数据表，已拒绝登记该文件。")
        if chart_count != expected_chart_count:
            raise DataWorkbookError("导出的 Excel 图表数量与受控合同不一致，已拒绝登记该文件。")

        overview = workbook["数据概览"]
        metric_values = [overview.cell(row=index, column=2).value for index in range(2, overview.max_row + 1)]
        if len(metric_values) < expected_metric_count:
            raise DataWorkbookError("导出的 Excel 缺少关键指标单元格，已拒绝登记该文件。")
        for expected, actual in zip(expected_metric_values, metric_values, strict=True):
            if not isinstance(actual, (int, float)) or not math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-12):
                raise DataWorkbookError("导出的 Excel 指标数值回读不一致，已拒绝登记该文件。")
        return DataWorkbookVerification(
            passed=True,
            sheet_names=names,
            table_count=table_count,
            chart_count=chart_count,
            metric_count=expected_metric_count,
        )
    finally:
        workbook.close()


def _next_output_path(directory: Path, dataset_name: str) -> Path:
    stem = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff_-]+", "_", Path(dataset_name).stem).strip("._") or "数据分析"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    candidate = directory / f"{stem}_分析结果_{timestamp}.xlsx"
    serial = 1
    while candidate.exists():
        candidate = directory / f"{stem}_分析结果_{timestamp}_{serial}.xlsx"
        serial += 1
    return candidate


def _safe_sheet_title(raw: str, existing: set[str]) -> str:
    base = re.sub(r"[\\/:*?\[\]]", "_", raw).strip() or "分析"
    base = base[:31]
    candidate = base
    serial = 2
    while candidate in existing:
        suffix = f"_{serial}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        serial += 1
    return candidate


def _column_letter(column_index: int) -> str:
    letters = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _finite_number(value: float) -> float:
    if not math.isfinite(value):
        raise DataWorkbookError("本次计算出现非有限数值，已拒绝导出。")
    return value


def _excel_cell(value: Any) -> Any:
    if value is None:
        return None
    if pd.isna(value):
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except ValueError:
            pass
    if isinstance(value, float):
        return _finite_number(value)
    return value
