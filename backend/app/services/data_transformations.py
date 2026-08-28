"""D5.3 的受限字段加工与新副本交付。

这里不是通用公式引擎。客户必须通过引导式选择构造有限的 ``DataTransformPlan``；所有数值、
日期和文本加工都由本地 pandas/openpyxl 完成，随后
重新打开新工作簿验证。这样模型不可用时仍能完成常见办公任务，源 CSV/XLSX 也始终不变。
"""

from __future__ import annotations

import math
import os
import re
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from app.core.config import settings
from app.schemas.data_agent import (
    DataColumnProfile,
    DataTransformFieldPreview,
    DataTransformOperationInput,
    DataTransformPlan,
    DataTransformPreviewRequest,
    DataTransformPreviewResponse,
    DataTransformationArtifact,
    DataTransformationExportRequest,
    DataTransformationExportResponse,
    DataTransformationVerification,
)
from app.services.data_workspace import DataWorkspaceError, load_data_dataset_for_analysis


_DEPENDENCIES_AVAILABLE = True
try:  # 与 D1-D3 一致：依赖问题返回可行动的客户错误，不让 API 变成 500。
    import pandas as pd
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - 正常 requirements 环境不会进入。
    pd = None
    _DEPENDENCIES_AVAILABLE = False


class DataTransformationError(ValueError):
    """字段加工计划、计算或新副本验证无法安全完成时抛出。"""


@dataclass(frozen=True)
class DataTransformationComputation:
    """预览和正式写入共用的一次确定性计算快照。"""

    preview: DataTransformPreviewResponse
    source_frame: Any
    transformed_frame: Any


_OPERATION_LABELS = {
    "arithmetic": "四则计算/比率",
    "date_part": "日期拆分",
    "round_number": "数值保留位数",
    "rank": "排名",
    "share": "占比",
    "segment": "分段标签",
    "cumulative": "累计",
    "period_change": "环比",
    "period_rate": "环比百分比",
    "text_trim": "文本轻清洗",
}


def preview_data_transformation(request: DataTransformPreviewRequest) -> DataTransformPreviewResponse:
    """仅在内存中计算并展示有限样例，绝不写文件。"""

    return compute_data_transformation(request).preview


def compute_data_transformation(request: DataTransformPreviewRequest) -> DataTransformationComputation:
    """从受控数据副本构造一组可审查的加工计划和派生字段。"""

    _ensure_dependencies()
    try:
        profile, source_frame = load_data_dataset_for_analysis(request.dataset_name)
    except DataWorkspaceError as exc:
        raise DataTransformationError(str(exc)) from exc

    if request.source_sha256.casefold() != profile.source_sha256.casefold():
        raise DataTransformationError("数据文件已变化或不属于当前预览，请重新选择材料后再加工。")

    transformed = source_frame.copy(deep=True)
    plans: list[DataTransformPlan] = []
    warnings: list[str] = []
    affected_count = 0
    empty_result_count = 0
    first_result: Any | None = None

    # 每项都在同一份内存副本上写入新列，但第一版只允许引用原始已画像字段；这样客户可以一次
    # 得到多列，仍不会把“连续公式链”变成难以复核的黑盒。`existing_columns` 随队列更新，确保
    # 新字段名不会互相覆盖。
    for ordinal, operation in enumerate(_operation_inputs(request), start=1):
        operation_request = request.model_copy(
            update={
                "operation_type": operation.operation_type,
                "primary_column": operation.primary_column,
                "secondary_column": operation.secondary_column,
                "result_column": operation.result_column,
                "date_part": operation.date_part,
                "arithmetic_operator": operation.arithmetic_operator,
                "round_digits": operation.round_digits,
                "operations": [],
            }
        )
        plan = build_data_transform_plan(
            operation_request,
            profile.columns,
            existing_columns={str(column) for column in transformed.columns},
            plan_ordinal=ordinal if ordinal > 1 else 0,
        )
        result, operation_warnings = _execute_transform(plan, transformed)
        transformed[plan.result_column] = result
        affected = int(result.notna().sum())
        empty = int(result.isna().sum())
        if empty:
            operation_warnings.append(f"新字段有 {empty} 行为空；已保留原行，未自动填补或删除。")
        warnings.extend(f"{plan.result_column}：{warning}" for warning in operation_warnings)
        plans.append(plan)
        affected_count += affected
        empty_result_count += empty
        if first_result is None:
            first_result = result

    if not plans or first_result is None:  # 防御性分支；Pydantic 默认允许空队列以兼容旧请求。
        raise DataTransformationError("字段加工队列为空，请至少选择一项可执行加工。")

    plan = plans[0]
    source_columns = [plan.primary_column]
    if plan.secondary_column:
        source_columns.append(plan.secondary_column)
    preview_rows: list[DataTransformFieldPreview] = []
    for row_index, (_, row) in enumerate(transformed.head(12).iterrows(), start=1):
        preview_rows.append(
            DataTransformFieldPreview(
                row_number=row_index,
                source_values=[_display_value(row[column]) for column in source_columns],
                result_value=_display_value(row[plan.result_column]),
            )
        )

    return DataTransformationComputation(
        preview=DataTransformPreviewResponse(
            plan=plan,
            plans=plans,
            row_count=len(transformed.index),
            affected_count=affected_count,
            empty_result_count=empty_result_count,
            previews=preview_rows,
            warnings=warnings[:12],
        ),
        source_frame=source_frame,
        transformed_frame=transformed,
    )


def _operation_inputs(request: DataTransformPreviewRequest) -> list[DataTransformOperationInput]:
    """统一兼容早期单项请求和新的一次多字段队列。"""

    if request.operations:
        return request.operations
    if request.operation_type is None:
        raise DataTransformationError("请在字段向导中选择一种加工方式，再选择对应字段。")
    return [
        DataTransformOperationInput(
            operation_type=request.operation_type,
            primary_column=request.primary_column,
            secondary_column=request.secondary_column,
            result_column=request.result_column,
            date_part=request.date_part,
            arithmetic_operator=request.arithmetic_operator,
            round_digits=request.round_digits,
        )
    ]


def export_data_transformation_workbook(
    request: DataTransformationExportRequest,
) -> DataTransformationExportResponse:
    """确认后写出字段加工副本，并重新打开核对工作表、行数和新字段。"""

    preview_request = DataTransformPreviewRequest(
        dataset_name=request.dataset_name,
        source_sha256=request.source_sha256,
        goal=request.goal,
        operation_type=request.operation_type,
        primary_column=request.primary_column,
        secondary_column=request.secondary_column,
        result_column=request.result_column,
        date_part=request.date_part,
        arithmetic_operator=request.arithmetic_operator,
        round_digits=request.round_digits,
        operations=request.operations,
    )
    computation = compute_data_transformation(preview_request)
    output_dir = settings.data_transformation_output_dir
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise DataTransformationError("无法创建字段加工输出目录，请检查磁盘权限和可用空间。") from exc

    destination = _next_output_path(output_dir, request.dataset_name)
    temporary = output_dir / f".{destination.stem}.{uuid.uuid4().hex}.partial{destination.suffix}"
    created_at = datetime.now(UTC).isoformat()
    try:
        _render_transformation_copy(temporary, computation)
        verification = _verify_transformation_copy(
            temporary,
            expected_result_columns=[plan.result_column for plan in computation.preview.plans],
            expected_row_count=len(computation.transformed_frame.index),
        )
        os.rename(temporary, destination)
    except DataTransformationError:
        raise
    except FileExistsError as exc:
        raise DataTransformationError("输出文件命名冲突，请重新确认加工。") from exc
    except OSError as exc:
        raise DataTransformationError("无法写入字段加工副本，请检查输出目录权限和可用空间。") from exc
    except Exception as exc:
        raise DataTransformationError("字段加工副本生成或回读验证失败，未生成正式交付文件。") from exc
    finally:
        try:
            temporary.unlink(missing_ok=True)
        except OSError:
            pass

    stat = destination.stat()
    preview = computation.preview
    return DataTransformationExportResponse(
        artifact=DataTransformationArtifact(
            name=destination.name,
            uri=f"agentflow-output://data_transformations/{destination.name}",
            size_bytes=stat.st_size,
            created_at=created_at,
        ),
        plan=preview.plan,
        plans=preview.plans,
        affected_count=preview.affected_count,
        empty_result_count=preview.empty_result_count,
        verification=verification,
        warnings=preview.warnings,
    )


def remove_data_transformation_output(result: DataTransformationExportResponse) -> None:
    """取消发生在渲染完成、artifact 登记之前时，删除未交付的新副本。"""

    candidate = (settings.data_transformation_output_dir / result.artifact.name).resolve()
    try:
        candidate.relative_to(settings.data_transformation_output_dir.resolve())
    except ValueError:
        return
    try:
        candidate.unlink(missing_ok=True)
    except OSError:
        # 取消任务已经在 SQLite 中稳定记录；无法立即清理时不把路径暴露给客户端。
        return


def build_data_transform_plan(
    request: DataTransformPreviewRequest,
    columns: list[DataColumnProfile],
    *,
    existing_columns: set[str] | None = None,
    plan_ordinal: int = 0,
) -> DataTransformPlan:
    """从有限操作和字段画像构造可审查计划；不能映射时明确拒绝。"""

    operation_type = request.operation_type
    if operation_type is None:
        raise DataTransformationError(
            "请在字段向导中选择一种加工方式，再选择对应字段。"
        )
    primary = _pick_required_column(columns, request.primary_column, _primary_kind(operation_type), "主字段")
    secondary: DataColumnProfile | None = None
    if operation_type == "arithmetic":
        secondary = _pick_required_column(columns, request.secondary_column, "number", "第二数值字段", exclude=primary.name)
    elif operation_type in {"cumulative", "period_change", "period_rate"}:
        secondary = _pick_required_column(columns, request.secondary_column, "date", "日期排序字段", exclude=primary.name)

    result_name = _safe_result_column_name(
        request.result_column or _default_result_name(operation_type, primary.name, secondary.name if secondary else None, request),
        existing=existing_columns or {column.name for column in columns},
    )
    parameters: dict[str, str | float | int | list[str]] = {}
    if operation_type == "arithmetic":
        parameters["operator"] = request.arithmetic_operator
    elif operation_type == "date_part":
        parameters["date_part"] = request.date_part
    elif operation_type == "round_number":
        parameters["digits"] = request.round_digits
    elif operation_type == "segment":
        parameters["strategy"] = "score_bands" if _looks_like_score(primary.name) else "equal_width_3"
    elif operation_type in {"cumulative", "period_change", "period_rate"}:
        parameters["sort"] = "ascending_date_then_original_row"
    scope = "全表；不分组。"
    if secondary and operation_type in {"cumulative", "period_change", "period_rate"}:
        scope = f"全表按“{secondary.name}”升序，日期相同则保持原始行顺序；不分组。"
    elif operation_type in {"rank", "share", "segment"}:
        scope = "全表范围；不按类别分组。"
    return DataTransformPlan(
        plan_id=(f"transform_{operation_type}_{plan_ordinal}" if plan_ordinal else f"transform_{operation_type}"),
        dataset_name=request.dataset_name,
        source_sha256=request.source_sha256,
        operation_type=operation_type,
        primary_column=primary.name,
        secondary_column=secondary.name if secondary else None,
        result_column=result_name,
        parameters=parameters,
        scope_description=scope,
        rationale=f"对“{primary.name}”执行{_OPERATION_LABELS[operation_type]}，结果写入新字段“{result_name}”。",
    )


def _execute_transform(plan: DataTransformPlan, frame: Any) -> tuple[Any, list[str]]:
    warnings: list[str] = []
    primary = frame[plan.primary_column]
    operation = plan.operation_type
    if operation == "text_trim":
        return primary.map(lambda value: value.strip() if isinstance(value, str) else value), warnings
    if operation == "date_part":
        parsed = pd.to_datetime(primary, errors="coerce")
        valid_count = int(parsed.notna().sum())
        if valid_count == 0 or valid_count / max(int(primary.notna().sum()), 1) < 0.8:
            raise DataTransformationError(f"“{plan.primary_column}”的有效日期比例不足 80%，不能安全拆分日期。")
        part = str(plan.parameters.get("date_part", "month"))
        if part == "year":
            return parsed.dt.year.astype("Int64"), warnings
        if part == "quarter":
            return parsed.map(lambda value: f"{value.year} Q{((value.month - 1) // 3) + 1}" if not pd.isna(value) else pd.NA), warnings
        if part == "weekday":
            labels = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
            return parsed.map(lambda value: labels[value.weekday()] if not pd.isna(value) else pd.NA), warnings
        return parsed.map(lambda value: f"{value.year}-{value.month:02d}" if not pd.isna(value) else pd.NA), warnings

    numeric = _numeric_series(primary)
    if int(numeric.notna().sum()) == 0:
        raise DataTransformationError(f"“{plan.primary_column}”没有可用于{_OPERATION_LABELS[operation]}的数值。")
    if operation == "arithmetic":
        assert plan.secondary_column is not None
        other = _numeric_series(frame[plan.secondary_column])
        if int(other.notna().sum()) == 0:
            raise DataTransformationError(f"“{plan.secondary_column}”没有可用于四则计算的数值。")
        operator = str(plan.parameters.get("operator", "multiply"))
        if operator == "add":
            return numeric + other, warnings
        if operator == "subtract":
            return numeric - other, warnings
        if operator == "divide":
            zero_count = int((other == 0).fillna(False).sum())
            if zero_count:
                warnings.append(f"“{plan.secondary_column}”有 {zero_count} 行为 0；除零结果已留空。")
            return numeric.div(other.where(other != 0)), warnings
        return numeric * other, warnings
    if operation == "round_number":
        return numeric.round(int(plan.parameters.get("digits", 2))), warnings
    if operation == "rank":
        return numeric.rank(method="min", ascending=False).astype("Int64"), warnings
    if operation == "share":
        total = float(numeric.sum(skipna=True))
        if math.isclose(total, 0.0, abs_tol=1e-12):
            raise DataTransformationError(f"“{plan.primary_column}”合计为 0，无法计算占比。")
        return numeric / total * 100.0, warnings
    if operation == "segment":
        return _segment_numeric_values(numeric, plan), warnings
    if operation in {"cumulative", "period_change", "period_rate"}:
        assert plan.secondary_column is not None
        dates = pd.to_datetime(frame[plan.secondary_column], errors="coerce")
        if int(dates.notna().sum()) == 0:
            raise DataTransformationError(f"“{plan.secondary_column}”没有可用于排序的有效日期。")
        order = dates.sort_values(kind="stable", na_position="last").index
        result = pd.Series(pd.NA, index=frame.index, dtype="Float64")
        ordered_values = numeric.loc[order]
        if operation == "cumulative":
            result.loc[order] = ordered_values.cumsum()
        elif operation == "period_change":
            result.loc[order] = ordered_values - ordered_values.shift(1)
            warnings.append("环比按相邻日期记录计算；日期相同的记录保持源文件顺序。")
        else:
            previous = ordered_values.shift(1)
            zero_count = int((previous == 0).fillna(False).sum())
            denominator = previous.abs().where(previous != 0)
            result.loc[order] = (ordered_values - previous).div(denominator) * 100.0
            if zero_count:
                warnings.append(f"上一期有 {zero_count} 行为 0；这些环比百分比已留空。")
            warnings.append("环比百分比按相邻日期记录计算；第一行没有上一期，会留空。")
        invalid_dates = int(dates.isna().sum())
        if invalid_dates:
            warnings.append(f"日期排序字段有 {invalid_dates} 行无法解析；这些行排在最后，结果可能为空。")
        return result, warnings
    raise DataTransformationError("当前加工操作不受支持。")


def _segment_numeric_values(values: Any, plan: DataTransformPlan) -> Any:
    finite = values.dropna()
    if finite.empty:
        return pd.Series(pd.NA, index=values.index, dtype="object")
    strategy = str(plan.parameters.get("strategy", "equal_width_3"))
    if strategy == "score_bands":
        return values.map(
            lambda value: pd.NA
            if pd.isna(value)
            else "优秀" if value >= 90 else "良好" if value >= 80 else "合格" if value >= 60 else "待提升"
        )
    minimum, maximum = float(finite.min()), float(finite.max())
    if math.isclose(minimum, maximum, abs_tol=1e-12):
        return values.map(lambda value: "同一数值" if not pd.isna(value) else pd.NA)
    lower = minimum + (maximum - minimum) / 3
    upper = minimum + (maximum - minimum) * 2 / 3
    plan.parameters["boundaries"] = [round(minimum, 6), round(lower, 6), round(upper, 6), round(maximum, 6)]
    return values.map(
        lambda value: pd.NA if pd.isna(value) else "低" if value <= lower else "中" if value <= upper else "高"
    )


def _primary_kind(operation_type: str) -> str:
    if operation_type == "date_part":
        return "date"
    if operation_type == "text_trim":
        return "text"
    return "number"


def _pick_required_column(
    columns: list[DataColumnProfile],
    requested: str | None,
    expected_kind: str,
    label: str,
    *,
    exclude: str = "",
) -> DataColumnProfile:
    candidates = [column for column in columns if column.name != exclude and _matches_kind(column, expected_kind)]
    if requested:
        selected = next((column for column in columns if column.name == requested), None)
        if selected is None:
            raise DataTransformationError(f"{label}“{requested}”不在当前数据中。")
        if selected.name == exclude or not _matches_kind(selected, expected_kind):
            raise DataTransformationError(f"{label}“{requested}”不适用于当前加工操作。")
        return selected
    if candidates:
        return candidates[0]
    required = "日期" if expected_kind == "date" else "数值" if expected_kind == "number" else "文本"
    raise DataTransformationError(f"当前数据未找到可用的{required}{label}，请在界面中选择其它操作或字段。")


def _matches_kind(column: DataColumnProfile, expected_kind: str) -> bool:
    if expected_kind == "number":
        return column.inferred_type == "number"
    if expected_kind == "date":
        return column.inferred_type == "date"
    return column.inferred_type in {"text", "mixed", "boolean"}


def _default_result_name(operation: str, primary: str, secondary: str | None, request: DataTransformPreviewRequest) -> str:
    if operation == "arithmetic":
        labels = {"add": "相加", "subtract": "相减", "multiply": "计算", "divide": "比率"}
        return f"{primary}_{labels[request.arithmetic_operator]}_{secondary or ''}".rstrip("_")
    if operation == "date_part":
        labels = {"year": "年份", "month": "月份", "quarter": "季度", "weekday": "星期"}
        return f"{primary}_{labels[request.date_part]}"
    if operation == "round_number":
        return f"{primary}_保留{request.round_digits}位"
    labels = {
        "rank": "排名",
        "share": "占比",
        "segment": "分档",
        "cumulative": "累计",
        "period_change": "环比",
        "period_rate": "环比百分比",
        "text_trim": "规范化",
    }
    return f"{primary}_{labels[operation]}"


def _safe_result_column_name(value: str, *, existing: set[str]) -> str:
    name = re.sub(r"[\r\n\t]+", " ", value).strip()
    if not name or len(name) > 180:
        raise DataTransformationError("新字段名称为空或过长，请使用 1 至 180 个字符的名称。")
    if name in existing:
        raise DataTransformationError(f"新字段“{name}”已存在；请换一个名称，避免覆盖原列。")
    return name


def _numeric_series(series: Any) -> Any:
    normalized = series.map(
        lambda value: str(value).replace(",", "").replace("¥", "").replace("￥", "").replace("$", "").strip()
        if value is not None and not pd.isna(value)
        else None
    )
    return pd.to_numeric(normalized, errors="coerce")


def _looks_like_score(column_name: str) -> bool:
    """只用稳定字段名判断成绩分段，避免自然语言悄悄改变副本计算规则。"""

    combined = column_name.casefold()
    return any(word in combined for word in ("成绩", "分数", "score", "考试"))


def _display_value(value: Any) -> str:
    if value is None or pd.isna(value):
        return "（空）"
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def _ensure_dependencies() -> None:
    if not _DEPENDENCIES_AVAILABLE:
        raise DataTransformationError("字段加工依赖未安装，请在 backend 目录安装 requirements.txt。")


def _render_transformation_copy(path: Path, computation: DataTransformationComputation) -> None:
    """按源文件类型写出一份干净的数据副本。

    交付物不是“分析报告工作簿”。它保留源数据原有的行列顺序，只在末尾追加客户确认的派生字段；
    CSV 仍是 CSV，Excel 仍是一张无额外样式的 Excel 表。加工说明和审计只保存在任务历史，避免
    普通客户打开文件后面对三张无关的说明/原始/加工工作表。
    """

    if path.suffix.casefold() == ".csv":
        computation.transformed_frame.to_csv(path, index=False, encoding="utf-8-sig")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据副本"
    _write_dataframe(sheet, computation.transformed_frame)
    workbook.save(path)
    workbook.close()


def _verify_transformation_copy(
    path: Path,
    *,
    expected_result_columns: list[str],
    expected_row_count: int,
) -> DataTransformationVerification:
    if path.suffix.casefold() == ".csv":
        try:
            copied = pd.read_csv(path, encoding="utf-8-sig")
        except Exception as exc:
            raise DataTransformationError("字段加工 CSV 副本无法重新读取验证。") from exc
        missing_columns = [column for column in expected_result_columns if column not in copied.columns]
        if missing_columns:
            raise DataTransformationError("字段加工副本未写入预期新字段，未生成正式交付文件。")
        if len(copied.index) != expected_row_count:
            raise DataTransformationError("字段加工副本行数与受控输入不一致，未生成正式交付文件。")
        return DataTransformationVerification(
            passed=True,
            sheet_names=[],
            row_count=len(copied.index),
            result_column=expected_result_columns[0] if expected_result_columns else "",
            result_columns=expected_result_columns,
        )
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        raise DataTransformationError("字段加工副本无法重新打开验证。") from exc
    try:
        expected_sheets = ["数据副本"]
        if workbook.sheetnames != expected_sheets:
            raise DataTransformationError("字段加工副本没有保持单表数据副本结构，未生成正式交付文件。")
        processed = workbook["数据副本"]
        headers = [processed.cell(row=1, column=index).value for index in range(1, processed.max_column + 1)]
        missing_columns = [column for column in expected_result_columns if column not in headers]
        if missing_columns:
            raise DataTransformationError("字段加工副本未写入预期新字段，未生成正式交付文件。")
        actual_row_count = max(processed.max_row - 1, 0)
        if actual_row_count != expected_row_count:
            raise DataTransformationError("字段加工副本行数与受控输入不一致，未生成正式交付文件。")
        return DataTransformationVerification(
            passed=True,
            sheet_names=expected_sheets,
            row_count=actual_row_count,
            result_column=expected_result_columns[0] if expected_result_columns else "",
            result_columns=expected_result_columns,
        )
    finally:
        workbook.close()


def _next_output_path(directory: Path, dataset_name: str) -> Path:
    safe_stem = re.sub(r"[\\/:*?\"<>|]+", "_", Path(dataset_name).stem).strip(" ._") or "数据集"
    prefix = f"{safe_stem}_字段加工"
    suffix = ".csv" if Path(dataset_name).suffix.casefold() == ".csv" else ".xlsx"
    candidate = directory / f"{prefix}{suffix}"
    if not candidate.exists():
        return candidate
    for ordinal in range(2, 10_000):
        candidate = directory / f"{prefix} ({ordinal}){suffix}"
        if not candidate.exists():
            return candidate
    raise DataTransformationError("同名字段加工副本过多，无法创建新的交付文件。")


def _write_dataframe(sheet: Any, frame: Any) -> None:
    sheet.append([str(column) for column in frame.columns])
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_excel_cell(value) for value in row])


def _excel_cell(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, float) and (not math.isfinite(value)):
        return None
    return value.item() if hasattr(value, "item") else value
